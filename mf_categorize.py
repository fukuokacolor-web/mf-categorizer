"""
mf_categorize.py ── マネーフォワード CSV を再分類して Excel レポートを出すスクリプト

使い方:
    python mf_categorize.py <CSVファイルパス>
    例:
        python mf_categorize.py ~/Downloads/収入・支出詳細_2026-05-01_2026-05-31.csv

出力:
    mf-categorizer/output/<元のCSV名>_categorized.xlsx
        ├ 取引明細   … 全取引(MFの分類 vs AIの分類を並べて表示。差異は黄色)
        ├ 月次サマリー … カテゴリ別合計
        └ 要確認     … ルールに無くて自動分類できなかった取引

学習のさせ方:
    rules.json に { "match": "店名の一部", "category": "...", "subcategory": "..." } を追加。
    次回からその店は自動で分類される。

必要ライブラリ:
    pip install openpyxl
"""

import csv
import json
import sys
from pathlib import Path

# Windows コンソール(cp932)でも絵文字が出せるよう stdout を utf-8 に再構成
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("⚠️ openpyxl が必要です。次のコマンドでインストール:")
    print("    pip install openpyxl")
    sys.exit(1)


SCRIPT_DIR = Path(__file__).parent
RULES_FILE = SCRIPT_DIR / "rules.json"
OUTPUT_DIR = SCRIPT_DIR / "output"

# 差異ハイライト用の色
COLOR_DIFF       = "FFF2CC"  # 薄い黄(MFとAIで分類が違うところ)
COLOR_UNKNOWN    = "F4CCCC"  # 薄い赤(要確認)
COLOR_TRANSFER   = "E0E0E0"  # 灰(振替は集計対象外)
COLOR_HEADER     = "B7DEE8"  # 薄い青


def load_rules():
    """rules.json を読む"""
    if not RULES_FILE.exists():
        print(f"⚠️ rules.json が見つかりません: {RULES_FILE}")
        return {"rules": [], "amount_rules": []}
    with open(RULES_FILE, encoding="utf-8") as f:
        return json.load(f)


def categorize(merchant: str, amount_abs: int, rules_data: dict):
    """1取引を分類する。

    戻り値: (category, subcategory, source)
        source: "amount_rule" | "rule" | "unknown"
    """
    m = merchant.strip()

    # 金額別ルールを先に確認(店名+金額の組み合わせ)
    for ar in rules_data.get("amount_rules", []):
        if ar["match"] in m:
            for split in ar["splits"]:
                max_amt = split.get("max_amount")
                if max_amt is None or amount_abs <= max_amt:
                    return (split["category"], split["subcategory"], "amount_rule")

    # 通常の店名ルール
    for r in rules_data.get("rules", []):
        if r["match"] in m:
            return (r["category"], r["subcategory"], "rule")

    return (None, None, "unknown")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    csv_path = Path(sys.argv[1])
    if not csv_path.exists():
        print(f"⚠️ ファイルが見つかりません: {csv_path}")
        sys.exit(1)

    rules_data = load_rules()
    n_rules = len(rules_data.get("rules", []))
    n_amount = len(rules_data.get("amount_rules", []))
    print(f"📖 ルール読込: 店名ルール {n_rules} 件 / 金額別ルール {n_amount} 件")

    # MF の CSV は cp932(Shift-JIS)
    rows = []
    with open(csv_path, encoding="cp932") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append(r)
    print(f"📥 取引読込: {len(rows)} 件")

    # 各取引を分類
    transactions = []
    summary = {}      # AIカテゴリ → 合計金額
    summary_mf = {}   # MFカテゴリ → 合計金額(比較用)
    unknowns = []

    for r in rows:
        date     = r.get("日付", "")
        merchant = r.get("内容", "")
        amount_s = r.get("金額（円）", "0")
        mf_cat   = r.get("大項目", "")
        mf_sub   = r.get("中項目", "")
        is_xfer  = r.get("振替", "0") == "1"

        try:
            amount = int(amount_s)
        except ValueError:
            amount = 0

        if is_xfer:
            ai_cat, ai_sub, src = ("振替", "振替", "transfer")
        else:
            ai_cat, ai_sub, src = categorize(merchant, abs(amount), rules_data)

        if ai_cat is None:
            # ルール未マッチ → MFの元分類で埋め、要確認に積む
            ai_cat = mf_cat or "未分類"
            ai_sub = mf_sub or "未分類"
            src = "unknown"
            unknowns.append({
                "date": date, "merchant": merchant, "amount": amount,
                "mf_cat": mf_cat, "mf_sub": mf_sub
            })

        transactions.append({
            "date": date, "merchant": merchant, "amount": amount,
            "mf_cat": mf_cat, "mf_sub": mf_sub,
            "ai_cat": ai_cat, "ai_sub": ai_sub,
            "source": src, "is_xfer": is_xfer
        })

        # サマリー(振替は除外)
        if not is_xfer:
            summary[ai_cat] = summary.get(ai_cat, 0) + amount
            summary_mf[mf_cat or "未分類"] = summary_mf.get(mf_cat or "未分類", 0) + amount

    # Excel 出力
    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = OUTPUT_DIR / f"{csv_path.stem}_categorized.xlsx"

    wb = Workbook()

    # ─── シート1: 取引明細 ───
    ws1 = wb.active
    ws1.title = "取引明細"
    headers = ["日付", "内容", "金額", "MF大項目", "MF中項目",
               "AI大項目", "AI中項目", "判定", "振替"]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)

    for t in transactions:
        ws1.append([
            t["date"], t["merchant"], t["amount"],
            t["mf_cat"], t["mf_sub"],
            t["ai_cat"], t["ai_sub"],
            t["source"], "○" if t["is_xfer"] else ""
        ])
        last_row = ws1.max_row
        # ハイライト
        if t["is_xfer"]:
            fill = PatternFill("solid", fgColor=COLOR_TRANSFER)
        elif t["source"] == "unknown":
            fill = PatternFill("solid", fgColor=COLOR_UNKNOWN)
        elif t["mf_cat"] != t["ai_cat"] or t["mf_sub"] != t["ai_sub"]:
            fill = PatternFill("solid", fgColor=COLOR_DIFF)
        else:
            fill = None
        if fill:
            for col in range(1, len(headers) + 1):
                ws1.cell(row=last_row, column=col).fill = fill

    widths = [12, 38, 12, 14, 14, 14, 14, 12, 6]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w
    ws1.freeze_panes = "A2"

    # ─── シート2: 月次サマリー ───
    ws2 = wb.create_sheet("月次サマリー")
    ws2.append(["カテゴリ(AI再分類)", "合計(円)", "件数"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)

    # AIカテゴリでの集計
    cat_count = {}
    for t in transactions:
        if not t["is_xfer"]:
            cat_count[t["ai_cat"]] = cat_count.get(t["ai_cat"], 0) + 1

    sorted_summary = sorted(summary.items(), key=lambda x: x[1])  # 支出は負なので小さい順 = 大きい支出から
    for cat, total in sorted_summary:
        ws2.append([cat, total, cat_count.get(cat, 0)])

    ws2.append([])
    ws2.append(["合計(振替除く)", sum(summary.values()),
                sum(1 for t in transactions if not t["is_xfer"])])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 8

    # ─── シート3: 要確認 ───
    ws3 = wb.create_sheet("要確認")
    ws3.append(["日付", "内容", "金額", "MFが付けた分類", "ヒント"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)

    # unknown_hints の適用
    hints = rules_data.get("unknown_hints", [])
    for u in unknowns:
        hint = ""
        for h in hints:
            if h["match"] in u["merchant"]:
                hint = h["hint"]
                break
        ws3.append([
            u["date"], u["merchant"], u["amount"],
            f"{u['mf_cat']} / {u['mf_sub']}".strip(" /"), hint
        ])

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 38
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 50

    # 保存
    wb.save(out_path)

    # ─── コンソール出力サマリー ───
    print()
    print("=" * 64)
    print("📊 再分類結果")
    print("=" * 64)
    n_total = len(transactions)
    n_xfer  = sum(1 for t in transactions if t["is_xfer"])
    n_rule  = sum(1 for t in transactions if t["source"] in ("rule", "amount_rule"))
    n_unk   = len(unknowns)
    n_diff  = sum(1 for t in transactions
                  if not t["is_xfer"]
                  and (t["mf_cat"] != t["ai_cat"] or t["mf_sub"] != t["ai_sub"]))
    print(f"  全取引            : {n_total} 件")
    print(f"  振替              : {n_xfer} 件 (集計対象外)")
    print(f"  ルールで自動分類  : {n_rule} 件")
    print(f"  要確認            : {n_unk} 件")
    print(f"  MFと判定が違うもの: {n_diff} 件 (黄色ハイライト)")
    print()
    print("【月次サマリー(支出は負、振替除く)】")
    for cat, total in sorted_summary:
        cnt = cat_count.get(cat, 0)
        print(f"  {cat:18s}  {total:>10,} 円  ({cnt} 件)")
    print(f"  {'─' * 18}  {'─' * 12}")
    print(f"  {'合計':18s}  {sum(summary.values()):>10,} 円")
    print()
    print(f"📁 出力: {out_path}")
    print()
    print("💡 要確認の取引を覚えさせるには rules.json に追記してください。")


if __name__ == "__main__":
    main()
